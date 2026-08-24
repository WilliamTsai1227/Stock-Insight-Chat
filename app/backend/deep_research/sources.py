"""
深度研究：使用者上傳檔案的前處理。

依副檔名分三條路，理由是 OpenAI File Search 的支援格式不含試算表，
而圖片本來就不該進向量庫：

  .pdf/.docx/.pptx/.txt/.md/... → 上傳到臨時 vector store，交給 FileSearchTool 檢索
  .xlsx/.xlsm/.csv             → 本地轉成 Markdown 表格，直接放進 prompt
  .png/.jpg/...                → 轉 base64 data URL，以 input_image 傳給多模態模型

vector store 只在這一次研究期間存在：研究結束（含失敗）立刻 `cleanup()` 刪除，
另外再加掛 expires_after 當保險絲，避免後端沒清乾淨時檔案留在 OpenAI 帳號裡。
"""

from __future__ import annotations

import asyncio
import base64
import csv
import io
import os
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException, UploadFile, status

from .config import (
    ACCEPTED_EXTENSIONS,
    IMAGE_EXTENSIONS,
    LEGACY_EXTENSION_HINTS,
    MAX_FILE_BYTES,
    MAX_FILE_MB,
    MAX_FILES,
    MAX_IMAGES,
    SPREADSHEET_EXTENSIONS,
    SPREADSHEET_MAX_CHARS,
    SPREADSHEET_TOTAL_MAX_CHARS,
    VECTOR_STORE_EXPIRES_DAYS,
)

_IMAGE_MEDIA_TYPES = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
    ".gif": "image/gif",
}

# 試算表轉 Markdown 的取樣上限（避免一張百萬列的表把 prompt 撐爆）
_SHEET_MAX_ROWS = 200
_SHEET_MAX_COLS = 30


def _ext_of(filename: str) -> str:
    return os.path.splitext(filename or "")[1].lower()


def _client():
    """深度研究專用的 AsyncOpenAI client（與 Agents SDK 共用同一把金鑰）。"""
    from openai import AsyncOpenAI

    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="伺服器未設定 OPENAI_API_KEY，深度研究無法使用。",
        )
    return AsyncOpenAI(api_key=api_key)


# ─────────────────────────────────────────────────────────────
# 讀檔與驗證
# ─────────────────────────────────────────────────────────────
@dataclass
class RawUpload:
    filename: str
    ext: str
    content: bytes

    @property
    def kind(self) -> str:
        if self.ext in IMAGE_EXTENSIONS:
            return "image"
        if self.ext in SPREADSHEET_EXTENSIONS:
            return "spreadsheet"
        return "document"


async def read_uploads(files: List[UploadFile]) -> List[RawUpload]:
    """把 UploadFile 讀進記憶體並做格式／大小驗證，任何一個不合格就整批拒絕。"""
    usable = [f for f in files if f and f.filename]
    if len(usable) > MAX_FILES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"一次最多上傳 {MAX_FILES} 個檔案（收到 {len(usable)} 個）。",
        )

    result: List[RawUpload] = []
    image_count = 0

    for upload in usable:
        ext = _ext_of(upload.filename)

        if ext in LEGACY_EXTENSION_HINTS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=LEGACY_EXTENSION_HINTS[ext],
            )
        if ext not in ACCEPTED_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"不支援的檔案格式：{upload.filename}。"
                    f"可接受 {'、'.join(sorted(ACCEPTED_EXTENSIONS))}。"
                ),
            )

        content = await upload.read()
        await upload.close()

        if not content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"檔案 {upload.filename} 是空的。",
            )
        if len(content) > MAX_FILE_BYTES:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"檔案 {upload.filename} 超過 {MAX_FILE_MB} MB 上限。",
            )

        if ext in IMAGE_EXTENSIONS:
            image_count += 1
            if image_count > MAX_IMAGES:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"一次最多上傳 {MAX_IMAGES} 張圖片。",
                )

        result.append(RawUpload(filename=upload.filename, ext=ext, content=content))

    return result


# ─────────────────────────────────────────────────────────────
# 試算表 → Markdown
# ─────────────────────────────────────────────────────────────
def _cells_to_markdown(rows: List[List[str]]) -> str:
    """第一列當表頭轉成 Markdown 表格；空表回空字串。"""
    if not rows:
        return ""

    # openpyxl 會把每列補滿到 max_col，尾端整欄皆空的要先砍掉，
    # 否則表格會拖著一串空欄位，白白吃掉 prompt 額度也難讀
    width = max(len(r) for r in rows)
    while width > 1 and all(not (r[width - 1].strip() if width <= len(r) else "") for r in rows):
        width -= 1

    padded = [(r + [""] * width)[:width] for r in rows]
    header, body = padded[0], padded[1:]

    def escape(cell: str) -> str:
        return (cell or "").replace("|", "\\|").replace("\n", " ").strip()

    lines = [
        "| " + " | ".join(escape(c) for c in header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    lines.extend("| " + " | ".join(escape(c) for c in r) + " |" for r in body)
    return "\n".join(lines)


def _xlsx_to_markdown(filename: str, content: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(
        io.BytesIO(content), read_only=True, data_only=True, keep_links=False
    )
    try:
        blocks: List[str] = []
        for sheet in workbook.worksheets:
            rows: List[List[str]] = []
            for row in sheet.iter_rows(
                max_row=_SHEET_MAX_ROWS, max_col=_SHEET_MAX_COLS, values_only=True
            ):
                if row is None or all(c is None for c in row):
                    continue
                rows.append(["" if c is None else str(c) for c in row])
            if not rows:
                continue
            truncated = (
                f"\n\n> （僅擷取前 {_SHEET_MAX_ROWS} 列）"
                if sheet.max_row and sheet.max_row > _SHEET_MAX_ROWS
                else ""
            )
            blocks.append(
                f"### 工作表：{sheet.title}\n\n{_cells_to_markdown(rows)}{truncated}"
            )
        return "\n\n".join(blocks)
    finally:
        workbook.close()


def _csv_to_markdown(content: bytes) -> str:
    text = content.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = [
        row for row in csv.reader(io.StringIO(text), dialect) if any(c.strip() for c in row)
    ]
    return _cells_to_markdown(rows[:_SHEET_MAX_ROWS])


def spreadsheet_to_markdown(upload: RawUpload) -> str:
    if upload.ext == ".csv":
        body = _csv_to_markdown(upload.content)
    else:
        body = _xlsx_to_markdown(upload.filename, upload.content)

    if len(body) > SPREADSHEET_MAX_CHARS:
        body = body[:SPREADSHEET_MAX_CHARS] + "\n\n> （內容過長，已截斷）"
    return body


# ─────────────────────────────────────────────────────────────
# 前處理結果
# ─────────────────────────────────────────────────────────────
@dataclass
class PreparedSources:
    names: List[str] = field(default_factory=list)
    vector_store_id: Optional[str] = None
    uploaded_file_ids: List[str] = field(default_factory=list)
    # (檔名, Markdown) —— 直接拼進 prompt
    spreadsheets: List[Tuple[str, str]] = field(default_factory=list)
    # (檔名, data URL) —— 以 input_image 傳給模型
    images: List[Tuple[str, str]] = field(default_factory=list)
    # 個別檔案處理失敗的說明（不中斷整體研究，只回報給前端）
    warnings: List[str] = field(default_factory=list)

    @property
    def has_file_search(self) -> bool:
        return bool(self.vector_store_id)

    def spreadsheet_prompt_block(self) -> str:
        """把試算表內容組成一段可直接附在使用者訊息後面的文字。"""
        if not self.spreadsheets:
            return ""

        chunks: List[str] = []
        budget = SPREADSHEET_TOTAL_MAX_CHARS
        for name, body in self.spreadsheets:
            if budget <= 0:
                chunks.append(f"## 附件試算表：{name}\n\n> （超出可讀取上限，已略過）")
                continue
            clipped = body[:budget]
            budget -= len(clipped)
            chunks.append(f"## 附件試算表：{name}\n\n{clipped}")

        return "以下是使用者上傳的試算表內容（已轉為 Markdown）：\n\n" + "\n\n".join(chunks)

    async def cleanup(self) -> None:
        """刪掉這次研究在 OpenAI 上留下的所有東西；失敗只記錄不拋出。"""
        if not self.vector_store_id and not self.uploaded_file_ids:
            return

        client = _client()
        try:
            if self.vector_store_id:
                try:
                    await client.vector_stores.delete(self.vector_store_id)
                except Exception as exc:  # noqa: BLE001 — 清理失敗不該影響已回傳的結果
                    print(
                        f"[DeepResearch] vector store 刪除失敗 "
                        f"{self.vector_store_id}: {type(exc).__name__}: {exc}",
                        flush=True,
                    )
                finally:
                    self.vector_store_id = None

            for file_id in self.uploaded_file_ids:
                try:
                    await client.files.delete(file_id)
                except Exception as exc:  # noqa: BLE001
                    print(
                        f"[DeepResearch] 檔案刪除失敗 {file_id}: "
                        f"{type(exc).__name__}: {exc}",
                        flush=True,
                    )
            self.uploaded_file_ids = []
        finally:
            await client.close()


async def prepare_sources(uploads: List[RawUpload]) -> PreparedSources:
    """
    把已驗證的上傳內容轉成 Agent 可用的三種形態。

    只要有任何一個文件類檔案，就會建立 vector store；失敗時會先把已建立的
    資源清掉再拋出，不會留下孤兒 vector store。
    """
    prepared = PreparedSources(names=[u.filename for u in uploads])

    for upload in uploads:
        if upload.kind == "image":
            media_type = _IMAGE_MEDIA_TYPES.get(upload.ext, "image/png")
            encoded = base64.b64encode(upload.content).decode("ascii")
            prepared.images.append(
                (upload.filename, f"data:{media_type};base64,{encoded}")
            )
        elif upload.kind == "spreadsheet":
            try:
                body = await asyncio.to_thread(spreadsheet_to_markdown, upload)
            except Exception as exc:  # noqa: BLE001 — 單一檔案壞掉不該中斷整場研究
                prepared.warnings.append(
                    f"{upload.filename} 解析失敗（{type(exc).__name__}），已略過。"
                )
                continue
            if body.strip():
                prepared.spreadsheets.append((upload.filename, body))
            else:
                prepared.warnings.append(f"{upload.filename} 沒有可讀取的內容，已略過。")

    documents = [u for u in uploads if u.kind == "document"]
    if not documents:
        return prepared

    client = _client()
    try:
        store = await client.vector_stores.create(
            name="deep-research-session",
            expires_after={
                "anchor": "last_active_at",
                "days": VECTOR_STORE_EXPIRES_DAYS,
            },
        )
        prepared.vector_store_id = store.id

        for doc in documents:
            vs_file = await client.vector_stores.files.upload_and_poll(
                vector_store_id=store.id,
                file=(doc.filename, doc.content),
            )
            prepared.uploaded_file_ids.append(vs_file.id)
            if vs_file.status != "completed":
                reason = getattr(getattr(vs_file, "last_error", None), "message", "")
                prepared.warnings.append(
                    f"{doc.filename} 建立索引失敗"
                    f"{f'（{reason}）' if reason else ''}，研究時無法檢索此檔。"
                )
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        await prepared.cleanup()
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"文件索引建立失敗：{type(exc).__name__}: {exc}",
        ) from exc
    finally:
        await client.close()

    # 全部文件都索引失敗時，掛著空的 vector store 只會讓 Agent 空轉
    if prepared.vector_store_id and len(prepared.warnings) >= len(documents):
        await prepared.cleanup()

    return prepared


def source_manifest(prepared: PreparedSources) -> List[Dict[str, Any]]:
    """回給前端顯示「這次讀了哪些檔案、用什麼方式讀」。"""
    manifest: List[Dict[str, Any]] = []
    doc_names = set(prepared.names) - {n for n, _ in prepared.spreadsheets} - {
        n for n, _ in prepared.images
    }
    for name in prepared.names:
        if name in {n for n, _ in prepared.images}:
            channel = "image"
        elif name in {n for n, _ in prepared.spreadsheets}:
            channel = "spreadsheet"
        elif name in doc_names:
            channel = "file_search"
        else:
            channel = "skipped"
        manifest.append({"name": name, "channel": channel})
    return manifest
